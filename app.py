from datetime import datetime, timedelta
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, login_required, logout_user, current_user
from werkzeug.security import generate_password_hash, check_password_hash
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill
from openpyxl.utils import get_column_letter
import io

app = Flask(__name__, instance_relative_config=True)
app.config['SECRET_KEY'] = os.urandom(24)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:////' + os.path.join(app.instance_path, 'workingtime.db')
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

# 确保实例文件夹存在
try:
    os.makedirs(app.instance_path)
except OSError:
    pass

db = SQLAlchemy(app)
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'

# 用户模型
class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(120), nullable=False)
    display_name = db.Column(db.String(80))
    timezone = db.Column(db.String(50), default='Asia/Shanghai')  # 默认使用中国时区
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    time_records = db.relationship('TimeRecord', backref='user', lazy=True)

    def get_local_time(self, utc_time=None):
        """将UTC时间转换为用户时区的时间"""
        from datetime import datetime
        import pytz
        
        if utc_time is None:
            utc_time = datetime.utcnow()
        
        if not isinstance(utc_time, datetime):
            return utc_time
            
        utc = pytz.UTC
        if utc_time.tzinfo is None:
            utc_time = utc.localize(utc_time)
            
        user_tz = pytz.timezone(self.timezone)
        return utc_time.astimezone(user_tz)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

# 工时记录模型
class TimeRecord(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    date = db.Column(db.Date, nullable=False)
    time_period = db.Column(db.String(20), nullable=False)  # 'morning' or 'afternoon'
    hours = db.Column(db.Float, nullable=False)
    project_name = db.Column(db.String(100))
    description = db.Column(db.Text)
    is_manual = db.Column(db.Boolean, default=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

def initialize_database():
    with app.app_context():
        db.create_all()
        # 检查是否需要添加 timezone 列
        from sqlalchemy import inspect
        inspector = inspect(db.engine)
        columns = [c['name'] for c in inspector.get_columns('user')]
        if 'timezone' not in columns:
            db.engine.execute('ALTER TABLE user ADD COLUMN timezone VARCHAR(50) DEFAULT "Asia/Shanghai"')
        
        user_count = User.query.count()
        if user_count == 0:
            admin_user = User(username='admin', display_name='管理员', timezone='Asia/Shanghai')
            admin_user.set_password('admin')
            db.session.add(admin_user)
            db.session.commit()

def needs_installation():
    with app.app_context():
        try:
            return User.query.count() == 0
        except:
            return True

# 路由定义
@app.route('/')
def index():
    if needs_installation():
        return redirect(url_for('install'))
    if not current_user.is_authenticated:
        return redirect(url_for('login'))
    return redirect(url_for('dashboard'))

@app.route('/install')
def install():
    if not needs_installation():
        return redirect(url_for('index'))
    return render_template('install.html')

@app.route('/install', methods=['POST'])
def do_install():
    if not needs_installation():
        return redirect(url_for('index'))
    initialize_database()
    flash('系统初始化成功！默认管理员账户: admin/admin', 'success')
    return redirect(url_for('install_success'))

@app.route('/install/success')
def install_success():
    return render_template('install_success.html')

@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        
        if user and user.check_password(password):
            login_user(user)
            return redirect(url_for('dashboard'))
        
        flash('用户名或密码错误', 'error')
    
    return render_template('login.html')

@app.route('/profile', methods=['GET', 'POST'])
@login_required
def profile():
    import pytz
    
    if request.method == 'POST':
        display_name = request.form.get('display_name')
        timezone = request.form.get('timezone')
        current_password = request.form.get('current_password')
        new_password = request.form.get('new_password')
        confirm_password = request.form.get('confirm_password')
        
        if new_password:
            if not current_user.check_password(current_password):
                flash('当前密码不正确', 'error')
                return redirect(url_for('profile'))
                
            if new_password != confirm_password:
                flash('新密码和确认密码不匹配', 'error')
                return redirect(url_for('profile'))
                
            current_user.set_password(new_password)
            flash('密码修改成功', 'success')
        
        if display_name:
            current_user.display_name = display_name
            flash('显示名称修改成功', 'success')
            
        if timezone and timezone in pytz.all_timezones:
            current_user.timezone = timezone
            flash('时区设置已更新', 'success')
        
        db.session.commit()
        return redirect(url_for('profile'))
    
    # 获取所有可用时区，但将用户所在区域的时区排在前面
    common_zones = [
        'Asia/Shanghai',  # 中国
        'Asia/Hong_Kong',
        'Asia/Taipei',
        'Asia/Tokyo',
        'Asia/Singapore',
        'US/Pacific',
        'US/Eastern',
        'Europe/London',
        'Europe/Paris'
    ]
    
    all_timezones = sorted(set(pytz.all_timezones) - set(common_zones))
    timezones = common_zones + all_timezones
        
    return render_template('profile.html', timezones=timezones)

@app.route('/logout')
@login_required
def logout():
    logout_user()
    return redirect(url_for('login'))

def calculate_total_hours(records):
    return sum(record.hours for record in records)

@app.route('/record/manual', methods=['GET', 'POST'])
@login_required
def add_manual_record():
    if request.method == 'GET':
        return render_template('record_time.html')
    
    # POST 处理
    date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    time_period = request.form['time_period']
    hours = float(request.form['hours'])
    project_name = request.form['project_name']
    description = request.form['description']
    next_page = request.form.get('next') or url_for('dashboard')

    record = TimeRecord(
        user_id=current_user.id,
        date=date,
        time_period=time_period,
        hours=hours,
        project_name=project_name,
        description=description,
        is_manual=True
    )
    
    db.session.add(record)
    db.session.commit()
    
    flash('工时记录已添加', 'success')
    return redirect(next_page)

@app.route('/dashboard')
@login_required
def dashboard():
    today = datetime.now().date()
    today_records = TimeRecord.query.filter_by(
        user_id=current_user.id,
        date=today
    ).all()
    
    # 计算今日上午和下午的工时
    morning_hours = sum(r.hours for r in today_records if r.time_period == 'morning')
    afternoon_hours = sum(r.hours for r in today_records if r.time_period == 'afternoon')
    total_hours = morning_hours + afternoon_hours
    
    return render_template('dashboard.html', 
                         today_records=today_records,
                         morning_hours=morning_hours,
                         afternoon_hours=afternoon_hours,
                         total_hours=total_hours,
                         current_time=datetime.now())

def generate_weekly_excel(user, week_dates, records):
    wb = Workbook()
    ws = wb.active
    
    # 设置基本样式
    thin_border = Border(left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')
    header_fill = PatternFill(start_color='E6E6FA', end_color='E6E6FA', fill_type='solid')
    bold_font = Font(bold=True)
    
    # 设置表头（包含姓名列，因此合并到AC列）
    start_date = week_dates[0]
    end_date = week_dates[6]
    title = f'周工时统计（{start_date.strftime("%m.%d")}-{end_date.strftime("%m.%d")}）'
    ws.merge_cells('A1:AC1')
    ws['A1'] = title
    ws['A1'].alignment = center_alignment
    ws['A1'].font = bold_font
    
    # 设置姓名列
    ws['A2'] = '姓名'
    ws.merge_cells('A2:A3')
    ws['A2'].alignment = center_alignment
    ws['A2'].font = bold_font
    ws['A2'].fill = header_fill
    ws.column_dimensions['A'].width = 15
    
    # 设置日期和时段标题
    for i, date in enumerate(week_dates):
        date_col = get_column_letter(i * 4 + 2)  # 从B列开始
        end_col = get_column_letter(i * 4 + 5)
        weekday_cn = ['(一)', '(二)', '(三)', '(四)', '(五)', '(六)', '(日)'][date.weekday()]
        ws[f'{date_col}2'] = f'{date.strftime("%m月%d日")}{weekday_cn}'
        ws.merge_cells(f'{date_col}2:{end_col}2')
        cell = ws[f'{date_col}2']
        cell.alignment = center_alignment
        cell.font = bold_font
        cell.fill = header_fill
        
        # 设置上午下午和工时列
        for subcol, (label, width) in enumerate([('上午', 15), ('工时', 8), ('下午', 15), ('工时', 8)]):
            col_letter = get_column_letter(i * 4 + 2 + subcol)
            cell = ws[f'{col_letter}3']
            cell.value = label
            cell.alignment = center_alignment
            cell.font = bold_font
            cell.fill = header_fill
            ws.column_dimensions[col_letter].width = width

    # 填充数据
    max_row = 4
    start_row = 4
    ws[f'A{start_row}'] = user.display_name or user.username
    
    for i, date in enumerate(week_dates):
        day_records = [r for r in records if r.date == date]
        morning_records = [r for r in day_records if r.time_period == 'morning']
        afternoon_records = [r for r in day_records if r.time_period == 'afternoon']
        
        base_col = i * 4 + 2  # 从B列开始
        current_row = start_row
        
        # 填充上午记录
        for record in morning_records:
            ws[f'{get_column_letter(base_col)}{current_row}'] = record.project_name
            ws[f'{get_column_letter(base_col + 1)}{current_row}'] = record.hours
            current_row += 1
        
        morning_end_row = current_row
        current_row = start_row
        
        # 填充下午记录
        for record in afternoon_records:
            ws[f'{get_column_letter(base_col + 2)}{current_row}'] = record.project_name
            ws[f'{get_column_letter(base_col + 3)}{current_row}'] = record.hours
            current_row += 1
        
        max_row = max(max_row, morning_end_row, current_row)
    
    # 合并姓名列
    if max_row > start_row:
        ws.merge_cells(f'A{start_row}:A{max_row-1}')
    
    # 设置所有单元格的边框和对齐方式
    for row in range(1, max_row):
        for col in range(1, 30):  # A到AC列
            cell = ws[f'{get_column_letter(col)}{row}']
            cell.border = thin_border
            cell.alignment = center_alignment
    
    # 设置行高
    for row in range(1, max_row):
        ws.row_dimensions[row].height = 20
    
    return wb

@app.route('/weekly')
@login_required
def weekly_view():
    # 如果是admin用户，重定向到admin的周视图
    if current_user.username == 'admin':
        return redirect(url_for('admin_weekly_view'))
        
    week_offset = request.args.get('week', type=int, default=0)
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]
    records = TimeRecord.query.filter(
        TimeRecord.user_id == current_user.id,
        TimeRecord.date.between(week_dates[0], week_dates[-1])
    ).all()
    
    # 计算每天的工时统计
    daily_stats = {}
    for date in week_dates:
        day_records = [r for r in records if r.date == date]
        morning_hours = sum(r.hours for r in day_records if r.time_period == 'morning')
        afternoon_hours = sum(r.hours for r in day_records if r.time_period == 'afternoon')
        daily_stats[date] = {
            'morning_hours': morning_hours,
            'afternoon_hours': afternoon_hours,
            'total_hours': morning_hours + afternoon_hours
        }
    
    # 计算本周总工时
    week_total = sum(stats['total_hours'] for stats in daily_stats.values())
    
    return render_template('weekly_view.html',
                         week_dates=week_dates,
                         records=records,
                         daily_stats=daily_stats,
                         week_total=week_total,
                         current_week_offset=week_offset)

@app.route('/weekly/export')
@login_required
def export_weekly():
    week_offset = request.args.get('week', type=int, default=0)
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]
    records = TimeRecord.query.filter(
        TimeRecord.user_id == current_user.id,
        TimeRecord.date.between(week_dates[0], week_dates[-1])
    ).all()
    
    # 生成Excel文件
    wb = generate_weekly_excel(current_user, week_dates, records)
    
    # 保存到内存中
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 生成文件名（加入用户姓名）
    user_name = current_user.display_name or current_user.username
    filename = f'{user_name}-工时统计_{start_of_week.strftime("%Y%m%d")}-{week_dates[6].strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/record', methods=['POST'])
@login_required
def add_record():
    date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
    time_period = request.form['time_period']
    hours = float(request.form['hours'])
    project_name = request.form['project_name']
    description = request.form['description']
    is_manual = True
    
    record = TimeRecord(
        user_id=current_user.id,
        date=date,
        time_period=time_period,
        hours=hours,
        project_name=project_name,
        description=description,
        is_manual=is_manual
    )
    
    db.session.add(record)
    db.session.commit()

@app.route('/record/<int:record_id>', methods=['GET', 'POST', 'DELETE'])
@login_required
def manage_record(record_id):
    record = TimeRecord.query.get_or_404(record_id)
    
    # 检查权限
    if record.user_id != current_user.id:
        return jsonify({'status': 'error', 'message': '无权访问此记录'}), 403
    
    if request.method == 'GET':
        return render_template('edit_record.html', record=record)
        
    elif request.method == 'POST':
        record.date = datetime.strptime(request.form['date'], '%Y-%m-%d').date()
        record.time_period = request.form['time_period']
        record.hours = float(request.form['hours'])
        record.project_name = request.form['project_name']
        record.description = request.form['description']
        db.session.commit()
        flash('工时记录已更新', 'success')
        return redirect(request.form.get('next') or url_for('dashboard'))
        
    elif request.method == 'DELETE':
        db.session.delete(record)
        db.session.commit()
        return jsonify({'status': 'success'})
    
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'status': 'success'})
    
    flash('工时记录已添加', 'success')
    return redirect(url_for('weekly_view'))

@app.route('/start_timer', methods=['POST'])
@login_required
def start_timer():
    import pytz
    utc_time = datetime.now(pytz.UTC)
    session['timer_start_time'] = utc_time.isoformat()
    session['project_name'] = request.form.get('project_name', '')
    return jsonify({'status': 'success'})

@app.route('/stop_timer', methods=['POST'])
@login_required
def stop_timer():
    if 'timer_start_time' not in session:
        return jsonify({'status': 'error', 'message': '计时器未启动'})
    
    import pytz
    user_tz = pytz.timezone(current_user.timezone)
    
    # 获取时间并确保时区信息正确
    start_time = datetime.fromisoformat(session['timer_start_time'])
    if start_time.tzinfo is None:
        start_time = pytz.UTC.localize(start_time)
    
    end_time = datetime.now(pytz.UTC)
    
    # 计算工时
    duration = end_time - start_time
    hours = round(duration.total_seconds() / 3600 * 2) / 2  # 四舍五入到最近的0.5小时
    hours = max(0.5, min(8.0, hours))  # 限制在0.5到8小时之间
    
    # 转换为用户时区来判断上午/下午
    local_time = end_time.astimezone(user_tz)
    time_period = 'morning' if local_time.hour < 12 else 'afternoon'
    project_name = session.get('project_name', '')
    
    # 创建工时记录
    # 获取补充说明
    data = request.get_json() if request.is_json else {}
    description = data.get('description', '')

    record = TimeRecord(
        user_id=current_user.id,
        date=local_time.date(),
        time_period=time_period,
        hours=hours,
        project_name=project_name,
        description=description,
        is_manual=False
    )
    
    db.session.add(record)
    db.session.commit()
    
    session.pop('timer_start_time', None)
    session.pop('project_name', None)
    
    return jsonify({
        'status': 'success',
        'hours': hours,
        'project_name': project_name
    })

@app.route('/users')
@login_required
def user_management():
    if current_user.username != 'admin':
        flash('权限不足', 'error')
        return redirect(url_for('dashboard'))
    
    users = User.query.all()
    return render_template('user_management.html', users=users)

@app.route('/users/add', methods=['POST'])
@login_required
def add_user():
    if current_user.username != 'admin':
        return jsonify({'status': 'error', 'message': '权限不足'})
    
    username = request.form.get('username')
    password = request.form.get('password')
    display_name = request.form.get('display_name')
    
    if User.query.filter_by(username=username).first():
        return jsonify({'status': 'error', 'message': '用户名已存在'})
    
    user = User(username=username, display_name=display_name)
    user.set_password(password)
    db.session.add(user)
    db.session.commit()
    
    return jsonify({'status': 'success'})

@app.route('/users/<int:user_id>', methods=['DELETE'])
@login_required
def delete_user(user_id):
    if current_user.username != 'admin':
        return jsonify({'status': 'error', 'message': '权限不足'})
    
    if current_user.id == user_id:
        return jsonify({'status': 'error', 'message': '不能删除自己'})
    
    user = User.query.get_or_404(user_id)
    db.session.delete(user)
    db.session.commit()
    
@app.route('/admin/weekly')
@login_required
def admin_weekly_view():
    if not current_user.username == 'admin':
        return redirect(url_for('weekly_view'))
    
    # 获取周偏移量和选中的用户
    week_offset = request.args.get('week', type=int, default=0)
    selected_user_id = request.args.get('user', type=int)
    
    if not selected_user_id:
        # 如果没有选择用户，默认显示第一个非admin用户
        first_user = User.query.filter(User.username != 'admin').first()
        if first_user:
            selected_user_id = first_user.id
        else:
            flash('没有可查看的用户', 'error')
            return redirect(url_for('dashboard'))
    
    # 获取选中的用户
    selected_user = User.query.get_or_404(selected_user_id)
    
    # 计算周起止日期
    today = datetime.now().date()
    start_of_week = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    week_dates = [start_of_week + timedelta(days=i) for i in range(7)]  # 周一到周日
    
    # 获取用户的记录
    records = TimeRecord.query.filter(
        TimeRecord.user_id == selected_user_id,
        TimeRecord.date.between(week_dates[0], week_dates[-1])
    ).order_by(TimeRecord.date, TimeRecord.time_period).all()
    
    # 计算每天的工时统计
    daily_stats = {}
    for date in week_dates:
        day_records = [r for r in records if r.date == date]
        morning_hours = sum(r.hours for r in day_records if r.time_period == 'morning')
        afternoon_hours = sum(r.hours for r in day_records if r.time_period == 'afternoon')
        daily_stats[date] = {
            'morning_hours': morning_hours,
            'afternoon_hours': afternoon_hours,
            'total_hours': morning_hours + afternoon_hours
        }
    
    # 计算本周总工时
    week_total = sum(stats['total_hours'] for stats in daily_stats.values())
    
    # 如果是 AJAX 请求，返回JSON数据（包含表格HTML和统计信息）
    if request.args.get('ajax'):
        try:
            # 打印请求信息
            print("AJAX请求参数:", dict(request.args))
            print("选中的用户:", selected_user.username if selected_user else 'None')
            
            # 渲染表格HTML
            table_html = render_template('admin_weekly_table.html',
                                       week_dates=week_dates,
                                       records=records,
                                       daily_stats=daily_stats,
                                       week_total=week_total,
                                       current_week_offset=week_offset,
                                       today=today,
                                       all_users=User.query.filter(User.username != 'admin'),
                                       selected_user=selected_user)
            
            # 打印渲染结果长度
            print("渲染的表格HTML长度:", len(table_html))
            
            # 构建响应数据
            response_data = {
                'table_html': table_html,
                'selected_user': selected_user.display_name or selected_user.username,
                'week_total': f"{week_total}",
                'avg_hours': f"{week_total / 5:.1f}",
                'debug_info': {
                    'request_args': dict(request.args),
                    'user_id': selected_user.id,
                    'week_offset': week_offset
                }
            }
            
            # 返回JSON数据
            response = jsonify(response_data)
            response.headers['Content-Type'] = 'application/json'
            print("发送JSON响应")
            return response
            
        except Exception as e:
            print("处理AJAX请求时出错:", str(e))
            # 返回错误信息
            return jsonify({
                'error': str(e),
                'status': 'error'
            }), 500

    return render_template('admin_weekly_view.html',
                         week_dates=week_dates,
                         records=records,
                         daily_stats=daily_stats,
                         week_total=week_total,
                         current_week_offset=week_offset,
                         today=today,
                         all_users=User.query.filter(User.username != 'admin'),
                         selected_user=selected_user)

@app.route('/admin/weekly/export/all')
@login_required
def admin_export_all_weekly():
    if not current_user.username == 'admin':
        return redirect(url_for('weekly_view'))
    
    week_offset = request.args.get('week', 0, type=int)
    
    # 计算周起止日期
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]  # 周一到周日
    
    # 获取所有非管理员用户
    users = User.query.filter(User.username != 'admin').order_by(User.username).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    
    # 创建汇总表
    summary_ws = wb.active
    summary_ws.title = "工时汇总"
    
    # 设置汇总表列宽
    summary_ws.column_dimensions['A'].width = 15  # 用户
    summary_ws.column_dimensions['B'].width = 10  # 周工时
    summary_ws.column_dimensions['C'].width = 10  # 日均工时
    
    # 汇总表样式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 写入汇总表表头
    summary_headers = ['用户', '周工时', '日均工时']
    for col, header in enumerate(summary_headers, 1):
        cell = summary_ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 记录每个用户的工时
    summary_row = 2
    
    # 为每个用户创建工作表并统计工时
    for user in users:
        # 获取用户本周记录
        records = TimeRecord.query.filter(
            TimeRecord.user_id == user.id,
            TimeRecord.date.between(week_dates[0], week_dates[-1])
        ).order_by(TimeRecord.date, TimeRecord.time_period).all()
        
        # 计算用户总工时
        total_hours = sum(record.hours for record in records)
        avg_hours = total_hours / 5 if total_hours > 0 else 0
        
        # 添加到汇总表
        summary_ws.cell(row=summary_row, column=1, value=user.display_name or user.username).border = border
        summary_ws.cell(row=summary_row, column=2, value=f"{total_hours}h").border = border
        summary_ws.cell(row=summary_row, column=3, value=f"{avg_hours:.1f}h").border = border
        
        # 设置单元格对齐方式
        for col in range(1, 4):
            cell = summary_ws.cell(row=summary_row, column=col)
            cell.alignment = Alignment(horizontal='center')
        
        summary_row += 1
        
        # 创建用户工作表
        ws = wb.create_sheet(title=f"{user.display_name or user.username}")
        
        # 设置列宽
        ws.column_dimensions['A'].width = 12  # 日期
        ws.column_dimensions['B'].width = 8   # 时段
        ws.column_dimensions['C'].width = 8   # 工时
        ws.column_dimensions['D'].width = 30  # 工作内容
        ws.column_dimensions['E'].width = 40  # 补充说明
        
        # 写入表头
        headers = ['日期', '时段', '工时', '工作内容', '补充说明']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.border = border
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 写入数据
        row = 2
        for record in records:
            ws.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d')).alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=2, value='上午' if record.time_period == 'morning' else '下午').alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=3, value=f"{record.hours}h").alignment = Alignment(horizontal='center')
            ws.cell(row=row, column=4, value=record.project_name)
            ws.cell(row=row, column=5, value=record.description or '')
            
            # 为每个单元格添加边框
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = border
            
            row += 1
    
    # 保存到内存
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'全员工时统计_{week_start.strftime("%Y%m%d")}.xlsx'
    )

@app.route('/admin/weekly/export')
@login_required
def admin_export_weekly():
    if not current_user.username == 'admin':
        return redirect(url_for('weekly_view'))
    
    week_offset = request.args.get('week', 0, type=int)
    selected_user_id = request.args.get('user', type=int)
    
    if not selected_user_id:
        flash('请选择要导出的用户', 'error')
        return redirect(url_for('admin_weekly_view'))
    
    # 获取选中的用户
    selected_user = User.query.get_or_404(selected_user_id)
    
    # 计算周起止日期
    today = datetime.now().date()
    week_start = today - timedelta(days=today.weekday()) + timedelta(weeks=week_offset)
    week_dates = [week_start + timedelta(days=i) for i in range(7)]  # 周一到周日
    
    # 获取记录
    records = TimeRecord.query.filter(
        TimeRecord.user_id == selected_user_id,
        TimeRecord.date.between(week_dates[0], week_dates[-1])
    ).order_by(TimeRecord.date, TimeRecord.time_period).all()
    
    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = f"{week_start.strftime('%Y-%m-%d')}周报表"
    
    # 设置列宽
    ws.column_dimensions['A'].width = 12  # 日期
    ws.column_dimensions['B'].width = 15  # 用户
    ws.column_dimensions['C'].width = 8   # 时段
    ws.column_dimensions['D'].width = 8   # 工时
    ws.column_dimensions['E'].width = 30  # 工作内容
    ws.column_dimensions['F'].width = 40  # 补充说明
    
    # 标题样式
    header_fill = PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                   top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 写入表头
    headers = ['日期', '用户', '时段', '工时', '工作内容', '补充说明']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.border = border
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 写入数据
    row = 2
    user_dict = {user.id: user for user in User.query.all()}
    
    for record in records:
        user = user_dict.get(record.user_id)
        if not user:
            continue
            
        ws.cell(row=row, column=1, value=record.date.strftime('%Y-%m-%d')).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=2, value=user.display_name or user.username).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=3, value='上午' if record.time_period == 'morning' else '下午').alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=4, value=record.hours).alignment = Alignment(horizontal='center')
        ws.cell(row=row, column=5, value=record.project_name)
        ws.cell(row=row, column=6, value=record.description or '')
        
        # 为每个单元格添加边框
        for col in range(1, 7):
            ws.cell(row=row, column=col).border = border
        
        row += 1
    
    # 在底部添加统计信息
    row += 1
    ws.cell(row=row, column=1, value='统计信息').font = Font(bold=True)
    row += 1
    
    # 计算每个用户的总工时
    user_totals = {}
    for record in records:
        user_totals[record.user_id] = user_totals.get(record.user_id, 0) + record.hours
    
    for user_id, total_hours in user_totals.items():
        user = user_dict.get(user_id)
        if user:
            ws.cell(row=row, column=1, value=user.display_name or user.username)
            ws.cell(row=row, column=2, value=f'{total_hours}小时')
            row += 1
    
    # 保存到内存
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # 生成文件名（包含用户姓名）
    user_name = selected_user.display_name or selected_user.username
    filename = f'{user_name}-工时统计_{week_start.strftime("%Y%m%d")}.xlsx'
    
    return send_file(
        excel_file,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

if __name__ == '__main__':
    app.run(debug=True)